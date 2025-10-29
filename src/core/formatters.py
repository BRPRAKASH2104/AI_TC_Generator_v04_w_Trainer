"""
Test case formatters for the AI Test Case Generator.

This module provides classes for formatting and outputting test cases to various formats,
with primary support for Excel/XLSX output with automotive-specific formatting.
"""

import json
from datetime import datetime
from typing import TYPE_CHECKING, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

if TYPE_CHECKING:
    from pathlib import Path

# Type aliases for better readability (PEP 695 style)
type TestCaseList = list[dict[str, Any]]
type FormattedOutput = dict[str, Any]


class TestCaseFormatter:
    """Formats test cases for output to Excel and other formats"""

    __slots__ = ("config", "logger")

    def __init__(self, config=None, logger=None):
        self.config = config
        self.logger = logger

    def format_to_excel(
        self, test_cases: TestCaseList, output_path: Path, metadata: dict[str, Any] | None = None
    ) -> bool:
        """
        Format test cases to Excel file with automotive-specific formatting.

        Args:
            test_cases: List of test case dictionaries
            output_path: Path where Excel file should be saved
            metadata: Additional metadata to include

        Returns:
            True if successful, False otherwise
        """
        try:
            # Prepare test case data
            formatted_cases = self._prepare_test_cases_for_excel(test_cases, metadata)

            if not formatted_cases:
                if self.logger:
                    self.logger.warning("No test cases to format")
                return False

            # Create DataFrame
            df = pd.DataFrame(formatted_cases)

            # Create Excel workbook with formatting
            self._create_formatted_excel(df, output_path, metadata)

            if self.logger:
                self.logger.info(f"Formatted {len(formatted_cases)} test cases to {output_path}")

            return True

        except Exception as e:
            if self.logger:
                self.logger.error(f"Error formatting to Excel: {e}")
            return False

    def _stringify_list(self, value: Any) -> str:
        """Converts a list of strings to a single newline-separated string."""
        if isinstance(value, list):
            return "\n".join(map(str, value))
        return str(value)  # Ensure the output is always a string

    def _prepare_test_cases_for_excel(
        self, test_cases: TestCaseList, metadata: dict[str, Any] | None = None
    ) -> list[dict[str, Any]]:
        """Prepare test cases with automotive-specific formatting (v03 style)"""
        formatted_cases = []

        # Get default values from config or use defaults
        default_values = self._get_default_test_values(metadata)

        for i, test_case in enumerate(test_cases, 1):
            requirement_id = test_case.get("requirement_id", "UNKNOWN")
            issue_id = self._generate_issue_id(test_case, i)

            # FIX: Support both v03 and v04 field names for backward compatibility
            # v04: summary, action, data, expected_result, summary_suffix
            # v03: feature_name, preconditions, test_steps, expected_result

            # Map v03 field names to v04 if present
            summary = (
                test_case.get("summary")
                or test_case.get("summary_suffix")
                or test_case.get("feature_name")
                or "Generated Test"
            )

            action = (
                test_case.get("action")
                or test_case.get("preconditions")
                or default_values["voltage_precondition"]
            )

            data_field = test_case.get("data") or test_case.get("test_steps") or "N/A"

            # v03 data formatting logic
            if isinstance(data_field, str) and data_field.startswith("1)"):
                data_field = (
                    data_field.replace(", ", "\n")
                    .replace("2)", "\n2)")
                    .replace("3)", "\n3)")
                    .replace("4)", "\n4)")
                    .replace("5)", "\n5)")
                )
            elif isinstance(data_field, list):
                data_field = "\n".join(str(step) for step in data_field)

            # FIX: Match v03 column structure exactly
            formatted_case = {
                "Issue ID": issue_id,
                "Summary": f"[{summary}]",  # v03 format: just [feature_name] without req_id
                "Test Type": default_values["test_type"],
                "Issue Type": default_values["issue_type"],
                "Project Key": default_values["project_key"],
                "Assignee": default_values["assignee"],
                "Description": "",  # Keep description empty as in v03
                "Action": self._stringify_list(action),
                "Data": data_field,
                "Expected Result": self._stringify_list(test_case.get("expected_result", "N/A")),
                "Planned Execution": default_values["planned_execution"],
                "Test Case Type": default_values["test_case_type"],
                "Feature Group": summary,  # v03 column: Feature Group
                "Components": default_values["components"],
                "Labels": self._stringify_list(default_values["labels"]),
                "LinkTest": requirement_id,  # v03 column name: LinkTest (not Tests)
            }

            formatted_cases.append(formatted_case)

        return formatted_cases

    def _get_default_test_values(self, metadata: dict[str, Any] = None) -> dict[str, Any]:
        """
        Get default values from config or use v03-compatible automotive defaults.

        FIX: Updated to match v03 static configuration exactly.
        """
        defaults = {
            "test_type": "RoboFit",  # v03: STATIC_TEST_TYPE
            "issue_type": "Test",  # v03: STATIC_ISSUE_TYPE
            "project_key": "TCTOIC",  # v03: STATIC_PROJECT_KEY
            "assignee": "ENGG",  # v03: STATIC_ASSIGNEE
            "planned_execution": "Manual",  # v03: STATIC_PLANNED_EXECUTION
            "test_case_type": "Feature Functional",  # v03: STATIC_TEST_CASE_TYPE
            "components": "Infotainment",  # v03: STATIC_COMPONENTS
            "labels": "SYS_DI_VALIDATION_TEST",  # v03: STATIC_LABELS
            "voltage_precondition": "1. Voltage= 12V\n2. Bat-ON",
        }

        # Override with config values if available
        if self.config and hasattr(self.config, "static_test"):
            config_test = self.config.static_test
            defaults.update(
                {
                    "issue_type": getattr(config_test, "issue_type", defaults["issue_type"]),
                    "project_key": getattr(config_test, "project_key", defaults["project_key"]),
                    "assignee": getattr(config_test, "assignee", defaults["assignee"]),
                    "test_case_type": getattr(
                        config_test, "test_case_type", defaults["test_case_type"]
                    ),
                    "planned_execution": getattr(
                        config_test, "planned_execution", defaults["planned_execution"]
                    ),
                    "components": getattr(config_test, "components", defaults["components"]),
                    "labels": getattr(config_test, "labels", defaults["labels"]),
                    "voltage_precondition": getattr(
                        config_test, "voltage_precondition", defaults["voltage_precondition"]
                    ),
                    "test_type": getattr(config_test, "test_type", defaults["test_type"]),
                }
            )

        # Override with metadata values if provided
        if metadata:
            defaults.update(metadata)

        return defaults

    def _generate_issue_id(self, test_case: dict[str, Any], index: int) -> str:
        """Generate unique issue ID for test case"""
        requirement_id = test_case.get("requirement_id", "UNKNOWN")
        test_id = test_case.get("test_id", f"TC_{index:03d}")

        # Return test_id if it already includes requirement_id, otherwise combine
        if requirement_id in test_id:
            return test_id
        else:
            return f"{requirement_id}_{test_id}"

    def _build_description(self, test_case: dict[str, Any]) -> str:
        """Build comprehensive test description"""
        parts = []

        # Add summary if available
        if test_case.get("summary"):
            parts.append(f"Summary: {self._stringify_list(test_case['summary'])}")

        # Add detailed action steps
        if test_case.get("action"):
            parts.append(f"Test Steps: {self._stringify_list(test_case['action'])}")

        # Add data requirements
        if test_case.get("data"):
            parts.append(f"Test Data: {self._stringify_list(test_case['data'])}")

        # Add expected results
        if test_case.get("expected_result"):
            parts.append(f"Expected Result: {self._stringify_list(test_case['expected_result'])}")

        # Add generation metadata
        if test_case.get("generation_time"):
            parts.append(f"Generated in {test_case['generation_time']:.2f}s")

        return "\n\n".join(parts) if parts else "AI-generated test case"

    def _create_formatted_excel(
        self, df: pd.DataFrame, output_path: Path, metadata: dict[str, Any] | None = None
    ) -> None:
        """Create Excel file with professional formatting"""
        # Ensure output directory exists
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # Add data to worksheet
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Apply formatting
        self._apply_excel_formatting(ws)

        # Add metadata sheet if provided
        if metadata:
            self._add_metadata_sheet(wb, metadata)

        # Save workbook
        wb.save(output_path)

    def _apply_excel_formatting(self, worksheet) -> None:
        """Apply professional formatting to Excel worksheet (v03 style)"""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Column widths (v03 style - 16 columns)
        column_widths = {
            "A": 15,  # Issue ID
            "B": 50,  # Summary
            "C": 15,  # Test Type
            "D": 12,  # Issue Type
            "E": 12,  # Project Key
            "F": 12,  # Assignee
            "G": 15,  # Description (empty)
            "H": 40,  # Action
            "I": 30,  # Data
            "J": 50,  # Expected Result
            "K": 15,  # Planned Execution
            "L": 20,  # Test Case Type
            "M": 30,  # Feature Group (v03 column)
            "N": 15,  # Components
            "O": 20,  # Labels
            "P": 20,  # LinkTest (v03 column name)
        }

        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        # Data row formatting
        data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = data_alignment

    def _add_metadata_sheet(self, workbook: Workbook, metadata: dict[str, Any]) -> None:
        """Add metadata information to separate sheet"""
        ws = workbook.create_sheet("Metadata")

        # Add generation info
        ws.append(["Generation Information", ""])
        ws.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append(["Model Used", metadata.get("model", "Unknown")])
        ws.append(["Template", metadata.get("template", "Default")])
        ws.append(["Total Test Cases", metadata.get("total_cases", 0)])

        # Add any additional metadata
        ws.append(["", ""])
        ws.append(["Additional Information", ""])
        for key, value in metadata.items():
            if key not in ["model", "template", "total_cases"]:
                ws.append([key, str(value)])

    def export_to_json(
        self, test_cases: TestCaseList, output_path: Path, metadata: dict[str, Any] | None = None
    ) -> bool:
        """Export test cases to JSON format"""
        try:
            output_data = {
                "metadata": metadata or {},
                "generated_at": datetime.now().isoformat(),
                "test_cases": test_cases,
                "total_count": len(test_cases),
            }

            output_path.parent.mkdir(parents=True, exist_ok=True)

            with output_path.open("w", encoding="utf-8") as f:
                json.dump(output_data, f, indent=2, ensure_ascii=False)

            if self.logger:
                self.logger.info(f"Exported {len(test_cases)} test cases to JSON: {output_path}")

            return True

        except Exception as e:
            if self.logger:
                self.logger.error(f"Error exporting to JSON: {e}")
            return False


class StreamingTestCaseFormatter(TestCaseFormatter):
    """Memory-efficient formatter for large test case sets"""

    def format_to_excel_streaming(
        self,
        test_cases_iterator,
        output_path: Path,
        metadata: dict[str, Any] | None = None,
        chunk_size: int = 100,
    ) -> bool:
        """Format test cases to Excel using streaming approach for memory efficiency"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Test Cases"

            # Write headers first (v03 style)
            headers = [
                "Issue ID",
                "Summary",
                "Test Type",
                "Issue Type",
                "Project Key",
                "Assignee",
                "Description",
                "Action",
                "Data",
                "Expected Result",
                "Planned Execution",
                "Test Case Type",
                "Components",
                "Labels",
                "Tests",
            ]
            ws.append(headers)

            # Apply header formatting
            self._apply_excel_formatting(ws)

            # Process test cases in chunks
            total_processed = 0
            chunk = []

            for test_case in test_cases_iterator:
                chunk.append(test_case)

                if len(chunk) >= chunk_size:
                    self._write_chunk_to_excel(ws, chunk, metadata, total_processed)
                    total_processed += len(chunk)
                    chunk = []

            # Write remaining chunk
            if chunk:
                self._write_chunk_to_excel(ws, chunk, metadata, total_processed)
                total_processed += len(chunk)

            # Add metadata sheet
            if metadata:
                metadata["total_cases"] = total_processed
                self._add_metadata_sheet(wb, metadata)

            # Save workbook
            output_path.parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)

            if self.logger:
                self.logger.info(f"Streamed {total_processed} test cases to {output_path}")

            return True

        except Exception as e:
            if self.logger:
                self.logger.error(f"Error in streaming format: {e}")
            return False

    def _write_chunk_to_excel(
        self, worksheet, chunk: TestCaseList, metadata: dict[str, Any], _start_index: int
    ) -> None:
        """Write a chunk of test cases to Excel worksheet (v03 style)"""
        formatted_chunk = self._prepare_test_cases_for_excel(chunk, metadata)

        for formatted_case in formatted_chunk:
            row_data = [
                formatted_case["Issue ID"],
                formatted_case["Summary"],
                formatted_case["Test Type"],
                formatted_case["Issue Type"],
                formatted_case["Project Key"],
                formatted_case["Assignee"],
                formatted_case["Description"],
                formatted_case["Action"],
                formatted_case["Data"],
                formatted_case["Expected Result"],
                formatted_case["Planned Execution"],
                formatted_case["Test Case Type"],
                formatted_case["Components"],
                formatted_case["Labels"],
                formatted_case["Tests"],
            ]
            worksheet.append(row_data)
